"""
Programing Assignment 2 - Superman
Math 248 Section 2
Members: Josh Bowers, Kendall Coleman, Jamie Fuller, William Westerkamp, and Fralin Widener

@author: Kendall Coleman
"""

# imports
import numpy as np
import openpyxl as op
from openpyxl import Workbook
from Altered_Newtons_Method import Altered_Newton_Method
from Fixed_point_iteration import fixedPointIteration1
from Fixed_point_iteration import fixedPointIteration2
from Halley_Method import Halley_Method
from Halley_Method import Halley_Method2
from Olvers_Method import Olver_Method
from Bisection import bisection
from Bisection import bisection2
from Bisection import bisection3
from newton_method_ex import Newton_Method_ex
from newton_method_ex import Newton_Method_ex2
from devious_function import f


def truncate(num: float, decimal: int) -> float:
    int_part, dec_part = str(num).split(".")
    return float(".".join((int_part, dec_part[:decimal])))


# make root array
root_array = []


# hard coded values
tol = 10**-14
b_tol = 1e-14

# run each value through newton's method and a fixed point method

# converges to -1.38750710558261
if (truncate(Newton_Method_ex(-2, tol), 14) == truncate(Olver_Method(-2, tol), 14)):
    root_1 = truncate(Newton_Method_ex(-2, tol), 14)
    print(f"This root converges to {root_1}.")

# converges to +-3.15145284329665
if (truncate(Newton_Method_ex2(4, tol), 14) == truncate(Halley_Method2(5, 10**-14), 14)):
    root_2 = truncate(Newton_Method_ex2(4, tol), 14)
    root_3 = -1 * truncate(Newton_Method_ex2(4, tol), 14)
    print(f"This root converges to {root_3} and {root_2}.")

# run each value through bisection method and a fixed point method

# converges to -0.51042934281782
if (truncate(bisection(-1, 0, b_tol, b_tol), 14) == truncate(fixedPointIteration1(-1, 10**-14), 14)):
    root_4 = truncate(bisection(-1, 0, b_tol, b_tol), 14)
    print(f"This root converges to {root_4}.")

# converges to +8.91069640296029
if (bisection(7, 9, b_tol, b_tol) == truncate(Halley_Method(9, 10**-14), 14)):
    root_5 = bisection(7, 9, b_tol, b_tol)
    print(f"This root converges to {root_5}.")

# converges to +-3.13108083857006
if (truncate(bisection2(3, 3.1478, b_tol, b_tol), 14) == truncate(fixedPointIteration2(3, 10**-14), 14)):
    root_6 = truncate(bisection2(3, 3.1478, b_tol, b_tol), 14)
    root_7 = -1 * truncate(bisection2(3, 3.1478, b_tol, b_tol), 14)
    print(f"This root converges to {root_7} and {root_6}.")

# converges to +0.5
if (bisection3(0, 1, b_tol, b_tol) == Altered_Newton_Method(1, tol)):
    root_8 = bisection3(0, 1, b_tol, b_tol)
    print(f"This root converges to {root_8}.")

# append to array in order

root_array.append(root_3)
root_array.append(root_7)
root_array.append(root_1)
root_array.append(root_4)
root_array.append(root_8)
root_array.append(root_6)
root_array.append(root_2)
root_array.append(root_5)
"""
root_array.append()
root_array.append()
"""

# create Lex_will_Perish

workbook = Workbook()
worksheet = workbook.active


root_index = 0  # counter
distance_a_array = []
distance_b_array = []
lock_box_codes = []
chopped_roots = []
for root in root_array:
    # set values

    r_k = root_array[root_index]

    # figure out lock box codes
    r_k_str = str(r_k)
    lock_box_str = ""
    if r_k == 0.5:
        lock_box_str = "000000"
    # special case 0.5
    else:
        for i in range(10, 16):
            if r_k_str[0] == "-":
                lock_box_str = lock_box_str + r_k_str[i + 1]
            else:
                lock_box_str = lock_box_str + r_k_str[i]
    lock_box_codes.append(lock_box_str)

    # chop off decimals 14 -> 4
    if r_k == 0.5:
        r_k = 0.5
    else:
        r_k = truncate(r_k, 4)
    chopped_roots.append(r_k)

    # find smaller distance
    if root_index == 0:
        distance = root_array[root_index + 1] - r_k
    elif root_index == len(root_array) - 1:
        distance = r_k - root_array[root_index - 1]
    else:
        if (root_array[root_index + 1] - r_k) > (r_k - root_array[root_index - 1]):
            distance = r_k - root_array[root_index - 1]
        else:
            distance = root_array[root_index + 1] - r_k

    a_k = distance * 0.1
    b_k = distance * 0.3

    distance_a_array.append(a_k)
    distance_b_array.append(b_k)

    # write into end_of_Lex excel file
    for i in range(len(chopped_roots)):
        worksheet.cell(row=i+1, column=1, value=chopped_roots[i])
        worksheet.cell(row=i+1, column=2, value=distance_a_array[i])
        worksheet.cell(row=i+1, column=3, value=distance_b_array[i])
    workbook.save("Lex_will_Perish.xlsx")
    # increment counter
    root_index = root_index + 1
print(root_array)
print(lock_box_codes)
