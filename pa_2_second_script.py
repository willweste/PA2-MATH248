"""
Programing Assignment 2 - Superman
Math 248 Section 2
Members: Josh Bowers, Kendall Coleman, Jamie Fuller, William Westerkamp, and Fralin Widener

@author: Kendall Coleman
@author: 

@version: 3.25.2024
"""
# imports
from op import Workbook
from devious_function import f
import numpy as np
import openpyxl as op

""" METHODS """


def bisection_method(a, b, iterations):
    """
    Bisection Method that finds root approximation after 3 iterations. 

    Parameters
    ----------
    a : float
        x_1 starting value.
    b : float
        x_2 ending value.
    iterations: int
        number of iterations to run

    Returns
    -------
    c : float
        final iterate.

    """
    x = np.zeros(iterations + 2)
    # set first value
    x[1] = (a + b)/2
    # set f(a) and f(c)
    f_a = f(a)
    f_x = f(x[1])

    for i in range(1, iterations + 1):
        if f_a*f_x < 0:
            # b = c
            b = x[i]
        else:
            # a = c
            a = x[i]
            f_a = f(a)

        # computes next values
        x[i+1] = (a + b) / 2
        f_x = f(x[i+1])

    # return final iterate
    return x[i]


def secant_method(a, b, iterations):
    """
    Secant Method that finds root approximation after 3 iterations.

    Parameters
    ----------
    a : float
        x_1 starting value.
    b : float
        x_2 ending value.
    iterations: int
        number of iterations to run

    Returns
    -------
    c : float
        final iterate.

    """
    x = np.zeros(iterations + 2)

    # set first values
    x[0] = a
    x[1] = b

    for i in range(1, iterations + 1):
        x[i + 1] = x[i] - (f(x[i]) * (x[i] - x[i-1])) / (f(x[i]) - f(x[i-1]))

    # return final iterate
    return x[i]


""" MAIN """


Lex_will_Perish = op.load_workbook("Lex_will_Perish.xlsx")
Lex_sheet = Lex_will_Perish.worksheets[0]

root_array = []
a_array = []
b_array = []

for column in Lex_sheet.iter_cols():
    column_title = column[0].value
    # check column name
    if column_title == "r_k":
        for cell in column:
            root_array.append(cell.value)
    elif column_title == "a_k":
        for cell in column:
            a_array.append(cell.value)
    elif column_title == "b_k":
        for cell in column:
            b_array.append(cell.value)

# create end_of_Lex

workbook = Workbook()
workbook.save(filename="end_of_Lex.xlsx")
worksheet = workbook.worksheets[0]

for root in root_array:
    # set values
    root_index = root

    r_k = root_array[root_index]
    a_k = a_array[root_index]
    b_k = b_array[root_index]

   # calculate starting points

    x_1 = r_k - a_k
    x_2 = r_k + b_k

    # run three iterations of Bisection Method and Secant Method and store final iterates

    bs_k = bisection_method(x_1, x_2, 3)
    sc_k = secant_method(x_1, x_2, 3)

    # write into end_of_Lex excel file
    r_k_cell = worksheet.cell(row=root_index, column=1)
    r_k_cell.value = r_k
    bs_k_cell = worksheet.cell(row=root_index, column=2)
    bs_k_cell.value = bs_k
    sc_k_cell = worksheet.cell(row=root_index, column=3)
    sc_k_cell.value = sc_k
